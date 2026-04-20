"""
2023 eXtra 월간 Sell-out → weekly_sellout DB 적재
- Source: 2023_Channel_Sell_Thru_Out_Stock_Consolidated.xlsx (Raw Data, Account='Extra')
- Mapping: Model_Mapping_Master_v6_Updated.xlsx (2023 Model Mapping)
- Month → 해당 월 15일 기준 ISO 주차 (mid-month 대표 주차)
"""
import sqlite3
import os
import sys
import openpyxl

SELLOUT_2023 = "/home/ubuntu/2026/10. Automation/01. Sell Out Dashboard/2023/2023_Channel_Sell_Thru_Out_Stock_Consolidated.xlsx"
MODEL_MAPPING = "/home/ubuntu/2026/07. Claude Rule/00. Model Mapping/Model_Mapping_Master_v6_Updated.xlsx"
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'sellout.db')

EXTRA_CHANNEL = 'United Electronics Company الشركة ا'

# 2023 month → mid-month ISO week
MONTH_TO_WEEK = {
    1: 'W2',  2: 'W7',  3: 'W11', 4: 'W15',
    5: 'W20', 6: 'W24', 7: 'W28', 8: 'W33',
    9: 'W37', 10: 'W41', 11: 'W46', 12: 'W50',
}

# SAP Category → fcst category label
CAT_MAP = {
    'Split AC':      'Inverter',
    'Window':        'Window AC',
    'Free Standing': 'Floor Standing AC',
}


def _build_model_map():
    wb = openpyxl.load_workbook(MODEL_MAPPING, read_only=True, data_only=True)
    ws = wb['2023 Model Mapping']
    rows = list(ws.iter_rows(values_only=True))
    mapping = {}
    for row in rows[1:]:   # skip header
        if not row or row[0] is None:
            continue
        _, group, model, mode, sap_cat, _, _, unified_code, compressor, _, _ = row
        if not unified_code:
            continue
        key = (str(group).strip(), _norm(str(model)), str(mode).strip())
        category = CAT_MAP.get(str(sap_cat).strip(), str(sap_cat).strip())
        mapping[key] = (str(unified_code).strip(), category)
    wb.close()
    return mapping


def _norm(s: str) -> str:
    """공백 정규화 (더블 스페이스 → 단일)."""
    return ' '.join(s.split())


def load(db_path: str = DB_PATH) -> int:
    model_map = _build_model_map()

    wb = openpyxl.load_workbook(SELLOUT_2023, read_only=True, data_only=True)
    ws = wb['Raw Data']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # header at row index 2, data from index 3
    data_rows = rows[3:]

    conn = sqlite3.connect(db_path)
    inserted = 0
    skipped = 0
    unmatched = set()

    for row in data_rows:
        if not row or row[0] is None:
            continue
        year, month, _, channel, account, group, model, mode, sell_thru, sell_out, stock = row
        if str(account).strip() != 'Extra':
            continue
        if not month or not sell_out:
            continue

        try:
            month_int = int(month)
            qty = float(sell_out)
        except (ValueError, TypeError):
            continue

        week = MONTH_TO_WEEK.get(month_int)
        if not week:
            continue

        key = (str(group).strip(), _norm(str(model)), str(mode).strip())
        if key not in model_map:
            unmatched.add(key)
            skipped += 1
            continue

        unified_code, category = model_map[key]

        try:
            conn.execute(
                "INSERT OR REPLACE INTO weekly_sellout "
                "(channel, year, week, model, category, qty, sellthru) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (EXTRA_CHANNEL, int(year), week, unified_code, category, qty,
                 float(sell_thru) if sell_thru else None)
            )
            inserted += 1
        except sqlite3.Error as e:
            print(f"  DB error: {e} | {key} → {unified_code}")

    conn.commit()
    conn.close()

    if unmatched:
        print(f"  Unmatched keys ({len(unmatched)}):")
        for k in sorted(unmatched):
            print(f"    {k}")

    return inserted


if __name__ == '__main__':
    n = load()
    print(f"Inserted {n} rows (2023 eXtra monthly → weekly_sellout)")
